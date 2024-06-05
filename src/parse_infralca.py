import logging
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from epd import LCAxEPD, LCAxProduct

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("INFRALCA")


def main(infralca_file: Path, output: Path):
    logger.info(f"Loading {infralca_file}")

    wb = load_workbook(filename=str(infralca_file), read_only=True, data_only=True)

    export_epds(wb["Emissionsfaktorer"], output / "epds")
    export_products(wb["Emissionsfaktorer"], output)

    logger.info("Done")


def export_epds(ws: Worksheet, output: Path):
    logger.info("Exporting EPDs")
    category = ""

    for row in ws.iter_rows(
            min_row=11, max_row=343, min_col=3, max_col=308, values_only=True
    ):
        if row[0] == "" or row[0] is None:
            continue
        elif row[0].startswith("-"):
            category = row[0].strip("-")
            logger.info(f"Parsing: {category}")
        else:
            logger.debug(f"Parsing row: {row[0]}")
            epd = LCAxEPD.from_row(
                row, category, ("3.1", date(year=2023, month=9, day=21))
            )
            logger.debug(f"Writing {epd.id} to file")
            (output / f"{epd.id}.json").write_text(epd.model_dump_json(indent=2))


def export_products(ws: Worksheet, output: Path):
    logger.info("Exporting Products")

    for row in ws.iter_rows(
            min_row=11, max_row=343, min_col=3, max_col=308, values_only=True
    ):
        if row[0] == "" or row[0] is None:
            continue
        elif row[0].startswith("-"):
            category = row[0].strip("-")
            logger.info(f"Parsing: {category}")
        else:
            logger.debug(f"Parsing row: {row[0]}")
            product = LCAxProduct.from_row(row, output)
            logger.debug(f"Writing {product.id} to file")
            (output / "products" / f"{product.id}.json").write_text(product.model_dump_json(indent=2))


if __name__ == "__main__":
    infralca_file = Path(__file__).parent / "infralca" / "InfraLCA_V3.1.xlsm"
    output_path = Path(__file__).parent.parent / "lcax"

    main(infralca_file, output_path)
